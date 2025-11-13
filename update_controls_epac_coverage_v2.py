#!/usr/bin/env python3
"""
Enhanced EPAC Coverage Analysis Script
- Accounts for Microsoft Defender for Cloud (MDC) implementation
- Identifies custom policy opportunities
- Comprehensive mapping of all controls

Usage:
    py update_controls_epac_coverage_v2.py

Requirements:
    py -m pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Enhanced control mapping with implementation details
CONTROL_MAPPING = {
    # ========================================
    # AU - AUDIT & ACCOUNTABILITY
    # ========================================
    "AU-2": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - Defender for Cloud monitors audit configurations",
        "implementation": "Azure Monitor, Log Analytics workspace, Diagnostic Settings policies enforce audit logging across all resources",
        "custom_policy_opportunity": "Create custom policy to enforce specific audit event types per organizational requirements",
        "priority": "IMPLEMENTED"
    },
    "AU-11": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors log retention settings",
        "implementation": "Log Analytics retention policies, Azure Storage immutability for long-term retention",
        "custom_policy_opportunity": "Create custom policy to enforce minimum 2-year retention as per control requirements",
        "priority": "IMPLEMENTED - CONSIDER CUSTOM RETENTION POLICY"
    },
    "AU-12": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC validates audit record generation",
        "implementation": "Built-in policies enforce diagnostic settings, audit logs enabled on all Azure resources",
        "custom_policy_opportunity": "N/A - Well covered by built-in policies",
        "priority": "IMPLEMENTED"
    },

    # ========================================
    # AC - ACCESS CONTROL
    # ========================================
    "AC-5": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors RBAC separation",
        "implementation": "Azure RBAC policies enforce role separation, Privileged Identity Management (PIM) for just-in-time access",
        "custom_policy_opportunity": "Create custom policy to audit separation of duties violations (e.g., same user with conflicting roles)",
        "priority": "IMPLEMENTED - CONSIDER CUSTOM AUDIT POLICY"
    },
    "AC-6": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC recommends least privilege violations",
        "implementation": "RBAC policies restrict permissions, MDC security score tracks over-privileged accounts",
        "custom_policy_opportunity": "Create deny policy for dangerous role assignments (Owner at subscription level)",
        "priority": "IMPLEMENTED - CUSTOM DENY POLICY RECOMMENDED"
    },
    "AC-6(1)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors privileged function access",
        "implementation": "Conditional Access policies, PIM for security functions, Azure AD roles for security administration",
        "custom_policy_opportunity": "N/A - Well covered",
        "priority": "IMPLEMENTED"
    },
    "AC-6(5)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC tracks privileged accounts",
        "implementation": "PIM enforces time-bound privileged access, MFA required for privileged roles",
        "custom_policy_opportunity": "Create custom policy to audit accounts with permanent privileged roles",
        "priority": "IMPLEMENTED - CUSTOM AUDIT RECOMMENDED"
    },
    "AC-6(7)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "PARTIAL - MDC provides data, manual review required",
        "implementation": "Azure AD access reviews, PIM audit logs",
        "custom_policy_opportunity": "Create automation to generate quarterly privilege review reports",
        "priority": "IMPLEMENTED - AUTOMATION OPPORTUNITY"
    },
    "AC-6(10)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors for privilege escalation",
        "implementation": "OS-level controls on VMs, Azure Policy prevents non-admin users from executing privileged functions",
        "custom_policy_opportunity": "Create guest configuration policy to audit OS-level privilege restrictions on VMs",
        "priority": "IMPLEMENTED - GUEST CONFIG POLICY RECOMMENDED"
    },
    "AC-18(3)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A - Physical environment control",
        "implementation": "Customer responsibility - Physical environment",
        "custom_policy_opportunity": "N/A - Cannot be enforced via Azure Policy",
        "priority": "CUSTOMER DOCUMENTATION REQUIRED"
    },
    "AC-21": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC can monitor data sharing",
        "implementation": "Customer responsibility - Information sharing agreements",
        "custom_policy_opportunity": "Create custom policy to enforce information classification tags, audit external sharing",
        "priority": "CUSTOM POLICY OPPORTUNITY - DATA CLASSIFICATION"
    },

    # ========================================
    # CA - SECURITY ASSESSMENT & AUTHORIZATION
    # ========================================
    "CA-3": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC can monitor network connections",
        "implementation": "Customer responsibility - Interconnection agreements documentation",
        "custom_policy_opportunity": "Create custom policy to enforce NSG rules, require approval workflow for ExpressRoute/VPN connections",
        "priority": "CUSTOM POLICY OPPORTUNITY - NETWORK GOVERNANCE"
    },

    # ========================================
    # CM - CONFIGURATION MANAGEMENT
    # ========================================
    "CM-2": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors configuration baselines",
        "implementation": "Azure Policy enforce baseline configurations, Azure Automanage for VM baselines",
        "custom_policy_opportunity": "Create custom baseline policies for organization-specific standards",
        "priority": "IMPLEMENTED - CUSTOM BASELINES RECOMMENDED"
    },
    "CM-2(2)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC uses Azure Policy for automation",
        "implementation": "Azure Policy Guest Configuration, Azure Automation State Configuration (DSC)",
        "custom_policy_opportunity": "Extend guest configuration policies for custom baseline requirements",
        "priority": "IMPLEMENTED - EXTEND AS NEEDED"
    },
    "CM-2(3)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "PARTIAL - Azure Repos/Git provides versioning",
        "implementation": "EPAC stores configs in Git, Azure Resource Manager tracks deployment history",
        "custom_policy_opportunity": "Create policy to enforce retention of ARM deployment history",
        "priority": "IMPLEMENTED - CONSIDER RETENTION POLICY"
    },
    "CM-3(2)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A - Process control",
        "implementation": "Customer responsibility - Change testing procedures in CI/CD",
        "custom_policy_opportunity": "Enforce via GitHub/GitLab branch protection, require approval gates",
        "priority": "IMPLEMENTED VIA CI/CD - DOCUMENTATION REQUIRED"
    },
    "CM-4": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC can assess security impact",
        "implementation": "Customer responsibility - Impact analysis procedures",
        "custom_policy_opportunity": "Create policy to require 'impactAssessment' tag on all change PRs",
        "priority": "CUSTOM POLICY OPPORTUNITY - CHANGE TRACKING"
    },
    "CM-4(2)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC validates post-change security posture",
        "implementation": "Customer responsibility - Control verification procedures",
        "custom_policy_opportunity": "Create automated testing policy - run MDC security assessment post-change",
        "priority": "AUTOMATION OPPORTUNITY - POST-CHANGE VALIDATION"
    },
    "CM-5": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC can audit unauthorized changes",
        "implementation": "Customer responsibility - Change control documentation",
        "custom_policy_opportunity": "Create policy to enforce Resource Locks on production resources, require break-glass approval",
        "priority": "CUSTOM POLICY OPPORTUNITY - RESOURCE LOCKS"
    },
    "CM-7": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC recommends disabling unnecessary services",
        "implementation": "Policies disable unused services, NSG rules enforce port restrictions",
        "custom_policy_opportunity": "Create deny policies for prohibited Azure services per organization requirements",
        "priority": "IMPLEMENTED - CUSTOM DENY LIST RECOMMENDED"
    },
    "CM-7(5)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC validates application control",
        "implementation": "Guest Configuration policies enforce application allowlists on VMs",
        "custom_policy_opportunity": "Create custom guest config policy for organization-approved software list",
        "priority": "IMPLEMENTED - CUSTOM ALLOWLIST RECOMMENDED"
    },
    "CM-8": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC provides asset inventory",
        "implementation": "Azure Resource Graph provides complete inventory, MDC asset inventory dashboard",
        "custom_policy_opportunity": "Create policy to enforce mandatory tags (Owner, CostCenter, Environment, DataClassification)",
        "priority": "IMPLEMENTED - CUSTOM TAG POLICY RECOMMENDED"
    },
    "CM-8(1)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC inventory updates in real-time",
        "implementation": "Azure Resource Graph auto-updates, Activity Log tracks changes",
        "custom_policy_opportunity": "N/A - Well covered",
        "priority": "IMPLEMENTED"
    },
    "CM-8(3)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC alerts on unauthorized resources",
        "implementation": "Azure Policy audits unauthorized resources, MDC security alerts",
        "custom_policy_opportunity": "Create deny policy for unapproved resource types/locations",
        "priority": "IMPLEMENTED - CUSTOM DENY POLICY RECOMMENDED"
    },
    "CM-9": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A - Documentation requirement",
        "implementation": "Customer responsibility - Configuration management plan",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "CM-12": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC shows resource locations",
        "implementation": "Customer responsibility - Data location documentation",
        "custom_policy_opportunity": "Create policy to enforce 'dataLocation' and 'dataClassification' tags, allowed regions policy",
        "priority": "CUSTOM POLICY OPPORTUNITY - DATA RESIDENCY"
    },
    "CM-12(1)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC Defender for Cloud Apps can discover data",
        "implementation": "Customer responsibility - Automated data discovery",
        "custom_policy_opportunity": "Use Microsoft Purview for data discovery, enforce classification labels",
        "priority": "INTEGRATION OPPORTUNITY - MICROSOFT PURVIEW"
    },

    # ========================================
    # CP - CONTINGENCY PLANNING (CRITICAL GAP!)
    # ========================================
    "CP-2": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A - Documentation requirement",
        "implementation": "Customer responsibility - Contingency plan documentation",
        "custom_policy_opportunity": "N/A - Documentation only, but can enforce Azure Site Recovery deployment",
        "priority": "CRITICAL - DOCUMENTATION + ASR POLICY"
    },
    "CP-2(3)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - RTO/RPO documentation",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "CRITICAL - DOCUMENTATION REQUIRED"
    },
    "CP-2(8)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Critical asset identification",
        "custom_policy_opportunity": "Create policy to require 'criticalityLevel' tag (High/Medium/Low)",
        "priority": "CUSTOM POLICY OPPORTUNITY - ASSET CRITICALITY TAGGING"
    },
    "CP-4": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - DR testing procedures",
        "custom_policy_opportunity": "N/A - Process documentation only",
        "priority": "CRITICAL - DOCUMENTATION + TESTING SCHEDULE"
    },
    "CP-4(1)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Coordinate DR testing",
        "custom_policy_opportunity": "N/A - Process documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "CP-6": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - Can validate geo-redundant storage",
        "implementation": "Customer responsibility - Alternate storage site",
        "custom_policy_opportunity": "Create policy to enforce GRS/GZRS storage accounts, require Azure Site Recovery",
        "priority": "CUSTOM POLICY OPPORTUNITY - GEO-REDUNDANCY"
    },
    "CP-6(1)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - Can audit paired regions",
        "implementation": "Customer responsibility - Storage separation",
        "custom_policy_opportunity": "Create policy to audit resources in paired regions, enforce multi-region deployment",
        "priority": "CUSTOM POLICY OPPORTUNITY - MULTI-REGION"
    },
    "CP-6(3)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Accessibility analysis",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "CP-7": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - Can validate multi-region deployment",
        "implementation": "Customer responsibility - Alternate processing site",
        "custom_policy_opportunity": "Create policy to require Traffic Manager/Front Door for multi-region apps",
        "priority": "CUSTOM POLICY OPPORTUNITY - MULTI-REGION APPS"
    },
    "CP-7(1)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - Can audit paired regions",
        "implementation": "Customer responsibility - Processing site separation",
        "custom_policy_opportunity": "Same as CP-6(1) - enforce multi-region",
        "priority": "CUSTOM POLICY OPPORTUNITY - MULTI-REGION"
    },
    "CP-7(2)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Processing accessibility",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "CP-7(3)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Priority of service",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "CP-9": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "YES - MDC recommends Azure Backup",
        "implementation": "Customer responsibility - Configure Azure Backup",
        "custom_policy_opportunity": "CREATE POLICY TO ENFORCE AZURE BACKUP ON ALL VMs AND DATABASES - HIGH PRIORITY!",
        "priority": "CRITICAL - CUSTOM BACKUP POLICY REQUIRED"
    },
    "CP-9(1)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC can monitor backup success",
        "implementation": "Customer responsibility - Test backups",
        "custom_policy_opportunity": "Create automation to schedule quarterly backup restore tests",
        "priority": "AUTOMATION OPPORTUNITY - BACKUP TESTING"
    },
    "CP-9(8)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "YES - Azure Backup encrypts at rest",
        "implementation": "Customer responsibility - Configure encryption",
        "custom_policy_opportunity": "Create policy to audit Azure Backup encryption settings",
        "priority": "CUSTOM AUDIT POLICY OPPORTUNITY"
    },
    "CP-10": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Recovery procedures",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "CRITICAL - DOCUMENTATION REQUIRED"
    },
    "CP-10(2)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Transaction recovery",
        "custom_policy_opportunity": "N/A - Application-level control",
        "priority": "APPLICATION DESIGN REQUIREMENT"
    },

    # ========================================
    # IA - IDENTIFICATION & AUTHENTICATION
    # ========================================
    "IA-12(5)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - Conditional Access can enforce",
        "implementation": "Customer responsibility - Out-of-band verification",
        "custom_policy_opportunity": "Enforce MFA with out-of-band methods via Conditional Access policy",
        "priority": "IMPLEMENTED VIA CONDITIONAL ACCESS"
    },

    # ========================================
    # PL - PLANNING
    # ========================================
    "PL-4": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Rules of behavior",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "PL-10": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Baseline selection",
        "custom_policy_opportunity": "N/A - Documentation only (document NIST 800-53 selection)",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "PL-11": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Baseline tailoring",
        "custom_policy_opportunity": "N/A - Documentation only",
        "priority": "DOCUMENTATION REQUIRED"
    },

    # ========================================
    # PS - PERSONNEL SECURITY
    # ========================================
    "PS-9": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Position descriptions",
        "custom_policy_opportunity": "N/A - HR documentation",
        "priority": "DOCUMENTATION REQUIRED"
    },

    # ========================================
    # RA - RISK ASSESSMENT
    # ========================================
    "RA-3": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC Secure Score provides risk insights",
        "implementation": "Customer responsibility - Formal risk assessment",
        "custom_policy_opportunity": "Use MDC Secure Score API to automate risk scoring, create custom risk register",
        "priority": "INTEGRATION OPPORTUNITY - MDC SECURE SCORE"
    },
    "RA-5(11)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC Defender for Cloud provides vulnerability reporting",
        "implementation": "MDC vulnerability assessment, public reporting channel can be configured",
        "custom_policy_opportunity": "N/A - Well covered by MDC",
        "priority": "IMPLEMENTED"
    },
    "RA-7": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC provides findings",
        "implementation": "Customer responsibility - Risk response procedures",
        "custom_policy_opportunity": "Create automation to assign MDC recommendations to teams, track remediation",
        "priority": "AUTOMATION OPPORTUNITY - RECOMMENDATION TRACKING"
    },
    "RA-9": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Criticality analysis",
        "custom_policy_opportunity": "Same as CP-2(8) - enforce criticality tagging",
        "priority": "CUSTOM POLICY OPPORTUNITY - CRITICALITY TAGS"
    },

    # ========================================
    # SA - SYSTEM & SERVICES ACQUISITION
    # ========================================
    "SA-4(9)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC shows open ports/protocols",
        "implementation": "Customer responsibility - Functions/ports documentation",
        "custom_policy_opportunity": "Create policy to require 'allowedPorts' tag, audit against NSG rules",
        "priority": "CUSTOM POLICY OPPORTUNITY - PORT DOCUMENTATION"
    },
    "SA-10": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Developer CM requirements",
        "custom_policy_opportunity": "N/A - Procurement/vendor management",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "SA-11": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Developer testing requirements",
        "custom_policy_opportunity": "N/A - Procurement/vendor management",
        "priority": "DOCUMENTATION REQUIRED"
    },
    "SA-15(3)": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "N/A",
        "implementation": "Customer responsibility - Criticality analysis requirements",
        "custom_policy_opportunity": "N/A - Procurement/vendor management",
        "priority": "DOCUMENTATION REQUIRED"
    },

    # ========================================
    # SC - SYSTEM & COMMUNICATIONS PROTECTION
    # ========================================
    "SC-2": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors function separation",
        "implementation": "Azure RBAC separates user and admin functions, managed identities",
        "custom_policy_opportunity": "N/A - Well covered",
        "priority": "IMPLEMENTED"
    },
    "SC-5": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC recommends DDoS Protection",
        "implementation": "Azure DDoS Protection Standard, Application Gateway WAF",
        "custom_policy_opportunity": "Create policy to require DDoS Protection on public IPs/VNETs",
        "priority": "IMPLEMENTED - CUSTOM ENFORCEMENT RECOMMENDED"
    },
    "SC-28": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC audits encryption at rest",
        "implementation": "Storage Service Encryption, Azure Disk Encryption, SQL TDE",
        "custom_policy_opportunity": "N/A - Well covered",
        "priority": "IMPLEMENTED"
    },
    "SC-28(1)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC validates encryption configurations",
        "implementation": "Policies enforce customer-managed keys (CMK), Azure Key Vault integration",
        "custom_policy_opportunity": "Create policy to enforce CMK usage (not Microsoft-managed keys)",
        "priority": "IMPLEMENTED - CUSTOM CMK POLICY RECOMMENDED"
    },
    "SC-39": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors container isolation",
        "implementation": "VM isolation, container isolation (AKS), network segmentation",
        "custom_policy_opportunity": "Create policy to enforce AKS network policies, private endpoints",
        "priority": "IMPLEMENTED - AKS POLICY RECOMMENDED"
    },

    # ========================================
    # SI - SYSTEM & INFORMATION INTEGRITY
    # ========================================
    "SI-2(2)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC Defender for Endpoint tracks patching",
        "implementation": "Azure Update Management, MDC vulnerability scanner",
        "custom_policy_opportunity": "Create policy to enforce update management enrollment on all VMs",
        "priority": "IMPLEMENTED - CUSTOM ENFORCEMENT RECOMMENDED"
    },
    "SI-4": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC is the primary monitoring tool",
        "implementation": "MDC security alerts, Azure Monitor, Log Analytics, Sentinel for SIEM",
        "custom_policy_opportunity": "N/A - Extremely well covered",
        "priority": "IMPLEMENTED"
    },
    "SI-4(1)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC provides system-wide IDS",
        "implementation": "MDC Defender for Servers, Network Security Groups, Azure Firewall threat intelligence",
        "custom_policy_opportunity": "N/A - Well covered",
        "priority": "IMPLEMENTED"
    },
    "SI-4(2)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC provides real-time analysis",
        "implementation": "MDC automated threat detection, Sentinel analytics rules",
        "custom_policy_opportunity": "Create custom Sentinel analytics rules for organization-specific threats",
        "priority": "IMPLEMENTED - CUSTOM SENTINEL RULES RECOMMENDED"
    },
    "SI-4(4)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC monitors network traffic",
        "implementation": "NSG flow logs, Azure Network Watcher, Traffic Analytics",
        "custom_policy_opportunity": "Create policy to enforce NSG flow logs on all subnets",
        "priority": "IMPLEMENTED - CUSTOM ENFORCEMENT RECOMMENDED"
    },
    "SI-7": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC File Integrity Monitoring",
        "implementation": "Azure Policy Guest Configuration, MDC File Integrity Monitoring (FIM)",
        "custom_policy_opportunity": "Extend FIM to cover additional file paths",
        "priority": "IMPLEMENTED - EXTEND AS NEEDED"
    },
    "SI-7(1)": {
        "covered": "YES",
        "epac_policy": "NIST 800-53 & MCSB",
        "mdc_coverage": "YES - MDC integrity checking",
        "implementation": "Trusted Launch for VMs, code signing validation",
        "custom_policy_opportunity": "Create policy to require Trusted Launch on all Gen2 VMs",
        "priority": "IMPLEMENTED - CUSTOM ENFORCEMENT RECOMMENDED"
    },
    "SI-10": {
        "covered": "PARTIAL",
        "epac_policy": "NIST 800-53 (partial)",
        "mdc_coverage": "PARTIAL - Application Gateway WAF provides some coverage",
        "implementation": "Application Gateway WAF, API Management policies",
        "custom_policy_opportunity": "Create policy to require WAF on all public web apps, enforce API Management",
        "priority": "CUSTOM POLICY OPPORTUNITY - WAF ENFORCEMENT"
    },

    # ========================================
    # SR - SUPPLY CHAIN RISK MANAGEMENT
    # ========================================
    "SR-10": {
        "covered": "NO",
        "epac_policy": "N/A",
        "mdc_coverage": "PARTIAL - MDC Defender for Containers scans images",
        "implementation": "Customer responsibility - Component inspection",
        "custom_policy_opportunity": "Create policy to require Defender for Containers, block unsigned images",
        "priority": "CUSTOM POLICY OPPORTUNITY - CONTAINER SECURITY"
    },
}

def get_coverage_summary():
    """Calculate coverage statistics"""
    total = len(CONTROL_MAPPING)
    yes = sum(1 for v in CONTROL_MAPPING.values() if v["covered"] == "YES")
    partial = sum(1 for v in CONTROL_MAPPING.values() if v["covered"] == "PARTIAL")
    no = sum(1 for v in CONTROL_MAPPING.values() if v["covered"] == "NO")

    custom_opportunities = sum(1 for v in CONTROL_MAPPING.values()
                              if "CUSTOM POLICY OPPORTUNITY" in v["priority"] or "CREATE POLICY" in v["custom_policy_opportunity"].upper())

    return {
        "total": total,
        "yes": yes,
        "partial": partial,
        "no": no,
        "custom_opportunities": custom_opportunities,
        "percentage_covered": round((yes + (partial * 0.5)) / total * 100, 1)
    }

def update_excel_file(file_path):
    """Update Excel with comprehensive EPAC coverage analysis"""
    print(f"Opening Excel file: {file_path}")

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        print(f"Working with sheet: {sheet.title}")

        # Find control column
        header_row = 1
        control_col = None

        for col in range(1, min(sheet.max_column + 1, 20)):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and any(keyword in str(cell_value).lower() for keyword in ['control', 'id', 'nist', 'ccaccr']):
                control_col = col
                print(f"Found control column at index {col}: {cell_value}")
                break

        if not control_col:
            print("Could not find control ID column. Using column 1.")
            control_col = 1

        # Add new columns
        base_col = sheet.max_column + 1
        epac_col = base_col
        mdc_col = base_col + 1
        impl_col = base_col + 2
        custom_col = base_col + 3
        priority_col = base_col + 4

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Add headers
        headers = [
            ("Managed by EPAC", epac_col),
            ("MDC Coverage", mdc_col),
            ("Implementation Details", impl_col),
            ("Custom Policy Opportunity", custom_col),
            ("Priority/Action", priority_col)
        ]

        for header_text, col_idx in headers:
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Process rows and add spacing
        rows_processed = 0
        rows_matched = 0
        import re

        # First pass: collect all control rows
        control_rows = []
        for row in range(header_row + 1, sheet.max_row + 1):
            control_cell = sheet.cell(row=row, column=control_col)
            control_id = str(control_cell.value).strip() if control_cell.value else ""
            if control_id:
                control_rows.append(row)

        # Second pass: process controls and insert blank rows
        rows_to_insert = []

        for row in control_rows:
            control_cell = sheet.cell(row=row, column=control_col)
            control_id = str(control_cell.value).strip() if control_cell.value else ""

            rows_processed += 1

            # Extract base control ID
            match = re.search(r'([A-Z]{2}-\d+(?:\(\d+\))?)', control_id)
            base_control = match.group(1) if match else None

            if base_control and base_control in CONTROL_MAPPING:
                rows_matched += 1
                info = CONTROL_MAPPING[base_control]

                # EPAC Coverage
                epac_cell = sheet.cell(row=row, column=epac_col)
                epac_cell.value = info["covered"]
                if info["covered"] == "YES":
                    epac_cell.fill = yes_fill
                    epac_cell.font = Font(bold=True, color="006100")
                elif info["covered"] == "PARTIAL":
                    epac_cell.fill = partial_fill
                    epac_cell.font = Font(bold=True, color="9C6500")
                else:
                    epac_cell.fill = no_fill
                    epac_cell.font = Font(bold=True, color="9C0006")
                epac_cell.alignment = Alignment(horizontal="center", vertical="center")

                # MDC Coverage
                mdc_cell = sheet.cell(row=row, column=mdc_col)
                mdc_cell.value = info["mdc_coverage"]
                mdc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Implementation
                impl_cell = sheet.cell(row=row, column=impl_col)
                impl_cell.value = info["implementation"]
                impl_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Custom Policy Opportunity
                custom_cell = sheet.cell(row=row, column=custom_col)
                custom_cell.value = info["custom_policy_opportunity"]
                custom_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if "HIGH PRIORITY" in info["custom_policy_opportunity"].upper() or "CRITICAL" in info["priority"]:
                    custom_cell.font = Font(bold=True, color="9C0006")

                # Priority
                priority_cell = sheet.cell(row=row, column=priority_col)
                priority_cell.value = info["priority"]
                priority_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if "CRITICAL" in info["priority"]:
                    priority_cell.fill = no_fill
                    priority_cell.font = Font(bold=True, color="9C0006")
                elif "CUSTOM POLICY OPPORTUNITY" in info["priority"]:
                    priority_cell.fill = partial_fill
                    priority_cell.font = Font(bold=True, color="9C6500")
            else:
                # Not mapped
                epac_cell = sheet.cell(row=row, column=epac_col)
                epac_cell.value = "NOT MAPPED"
                epac_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Mark this row for adding a blank row after it
            rows_to_insert.append(row)

        # Third pass: insert blank rows (work backwards to avoid shifting issues)
        print(f"\nAdding blank rows for readability...")
        for row in reversed(rows_to_insert):
            sheet.insert_rows(row + 1, 1)
            # Set row height for blank row to be smaller
            sheet.row_dimensions[row + 1].height = 6

        # Adjust column widths
        sheet.column_dimensions[openpyxl.utils.get_column_letter(epac_col)].width = 18
        sheet.column_dimensions[openpyxl.utils.get_column_letter(mdc_col)].width = 35
        sheet.column_dimensions[openpyxl.utils.get_column_letter(impl_col)].width = 50
        sheet.column_dimensions[openpyxl.utils.get_column_letter(custom_col)].width = 50
        sheet.column_dimensions[openpyxl.utils.get_column_letter(priority_col)].width = 35

        # Create summary sheet
        stats = get_coverage_summary()
        summary_sheet = workbook.create_sheet(title="EPAC Coverage Summary", index=0)

        summary_data = [
            ["EPAC & MDC Coverage Analysis", ""],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Overall Statistics", ""],
            ["Total Controls Analyzed:", stats["total"]],
            ["Fully Covered (YES):", f"{stats['yes']} ({round(stats['yes']/stats['total']*100, 1)}%)"],
            ["Partially Covered (PARTIAL):", f"{stats['partial']} ({round(stats['partial']/stats['total']*100, 1)}%)"],
            ["Not Covered (NO):", f"{stats['no']} ({round(stats['no']/stats['total']*100, 1)}%)"],
            ["Overall Coverage Score:", f"{stats['percentage_covered']}%"],
            ["", ""],
            ["Custom Policy Opportunities", ""],
            ["Controls with Custom Policy Opportunities:", stats["custom_opportunities"]],
            ["", ""],
            ["Priority Actions", ""],
            ["1. CP-9: CREATE BACKUP ENFORCEMENT POLICY", "CRITICAL - Require Azure Backup on all VMs/DBs"],
            ["2. CP Controls: Document DR/BCP procedures", "CRITICAL - 0% coverage for contingency planning"],
            ["3. CM-8/CM-12: Enforce mandatory tagging", "HIGH - Data classification, criticality, owner tags"],
            ["4. SC-5: Enforce DDoS Protection", "HIGH - Require on public IPs/VNETs"],
            ["5. SI-2(2): Enforce Update Management", "HIGH - All VMs must enroll"],
            ["", ""],
            ["Quick Wins (Custom Policies)", ""],
            ["• Mandatory Tags: Owner, CostCenter, Environment, DataClassification, CriticalityLevel", ""],
            ["• Backup Enforcement: Azure Backup required on all VMs and SQL databases", ""],
            ["• Geo-Redundancy: Require GRS/GZRS storage, multi-region for critical apps", ""],
            ["• DDoS Protection: Require Standard tier on public-facing resources", ""],
            ["• Update Management: Require enrollment on all compute resources", ""],
            ["• Resource Locks: Require locks on production resource groups", ""],
            ["• WAF Enforcement: Require Application Gateway WAF on public web apps", ""],
        ]

        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16, color="0066CC")
                elif row_data[0] in ["Overall Statistics", "Custom Policy Opportunities", "Priority Actions", "Quick Wins (Custom Policies)"]:
                    cell.font = Font(bold=True, size=12)
                elif "CRITICAL" in str(value):
                    cell.font = Font(bold=True, color="9C0006")
                elif "HIGH" in str(value):
                    cell.font = Font(bold=True, color="9C6500")

        summary_sheet.column_dimensions['A'].width = 60
        summary_sheet.column_dimensions['B'].width = 40

        # Save
        backup_path = file_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        print(f"\nCreating backup: {backup_path}")
        workbook.save(backup_path)

        print(f"Saving updated file: {file_path}")
        workbook.save(file_path)

        print(f"\n✅ SUCCESS!")
        print(f"   - Processed {rows_processed} rows")
        print(f"   - Matched {rows_matched} controls")
        print(f"   - Coverage Score: {stats['percentage_covered']}%")
        print(f"   - Custom Policy Opportunities: {stats['custom_opportunities']}")

        return True

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_file = "Azure_CC_Local_Controls.xlsx"

    print("=" * 80)
    print("Enhanced EPAC & MDC Coverage Analysis Tool")
    print("=" * 80)
    print()

    success = update_excel_file(excel_file)

    if success:
        stats = get_coverage_summary()
        print("\n" + "=" * 80)
        print("Key Findings:")
        print("=" * 80)
        print(f"✅ {stats['yes']} controls fully managed by EPAC/MDC")
        print(f"⚠️  {stats['partial']} controls partially covered")
        print(f"❌ {stats['no']} controls require manual procedures")
        print(f"🔧 {stats['custom_opportunities']} custom policy opportunities identified")
        print(f"\n📊 Overall Coverage: {stats['percentage_covered']}%")
        print("\nCheck 'EPAC Coverage Summary' sheet for priority actions!")
    else:
        print("\nPlease check errors above.")
